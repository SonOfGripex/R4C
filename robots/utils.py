import io
from openpyxl import Workbook


def create_robots_excel(robots):
    workbook = Workbook()

    robots_by_model = {}
    for robot in robots:
        if robot.model not in robots_by_model:
            robots_by_model[robot.model] = []
        robots_by_model[robot.model].append(robot)

    for model, robots_of_model in robots_by_model.items():
        worksheet = workbook.create_sheet(title=f"Модель {model}")

        worksheet.append(["Версия", "Количество"])

        robots_by_version = {}
        for robot in robots_of_model:
            if robot.version not in robots_by_version:
                robots_by_version[robot.version] = []
            robots_by_version[robot.version].append(robot)

        sorted_versions = sorted(robots_by_version.keys())
        for version in sorted_versions:
            worksheet.append([version, len(robots_by_version[version])])

    if "Sheet" in workbook.sheetnames:
         std=workbook["Sheet"]
         workbook.remove(std)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()